from flask import Flask, render_template, request, redirect, flash, send_file, send_from_directory
from config import Config
from database import (
    db,
    Moto,
    MotoImportada,
    Leitura,
    HistoricoMoto
)
from utils.exportador import exportar_txt
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from io import BytesIO
import csv
import io
import os


# ============================================================
# CAMINHO DA PASTA RAIZ DO PROJETO
# ============================================================

BASE_DIR = os.path.dirname(
    os.path.abspath(__file__)
)


# ============================================================
# CONFIGURAÇÃO DO FLASK
# OS HTML ESTÃO NA PASTA RAIZ
# ============================================================

app = Flask(
    __name__,
    template_folder=BASE_DIR,
    static_folder=None
)

app.config.from_object(Config)

app.secret_key = "rfidmotos123"

db.init_app(app)


# ============================================================
# CRIA AS TABELAS
# NÃO APAGA DADOS EXISTENTES
# ============================================================

with app.app_context():
    db.create_all()


# ============================================================
# CSS
# O style.css ESTÁ NA PASTA RAIZ
# ============================================================

@app.route("/style.css")
def style_css():

    return send_from_directory(
        BASE_DIR,
        "style.css",
        mimetype="text/css"
    )


# ============================================================
# PÁGINA INICIAL
# ============================================================

@app.route("/")
def index():

    motos = Moto.query.order_by(
        Moto.id.desc()
    ).all()

    return render_template(
        "index.html",
        motos=motos
    )


# ============================================================
# CADASTRAR MOTO
# ============================================================

@app.route("/cadastrar", methods=["GET", "POST"])
def cadastrar():

    if request.method == "POST":

        # ----------------------------------------------------
        # DADOS DO FORMULÁRIO
        # ----------------------------------------------------

        tag_texto = request.form.get(
            "tag",
            ""
        ).strip()

        chassi = request.form.get(
            "chassi",
            ""
        ).strip()

        modelo_digitado = request.form.get(
            "modelo",
            ""
        ).strip()

        cor_digitada = request.form.get(
            "cor",
            ""
        ).strip()

        ano_digitado = request.form.get(
            "ano",
            ""
        ).strip()

        placa = request.form.get(
            "placa",
            ""
        ).strip()

        montador = request.form.get(
            "montador",
            ""
        ).strip()


        # ----------------------------------------------------
        # VERIFICA TAG
        # ----------------------------------------------------

        try:

            tag = int(tag_texto)

        except ValueError:

            flash(
                "A TAG RFID deve ser um número.",
                "erro"
            )

            return redirect("/cadastrar")


        if tag < 1 or tag > 10:

            flash(
                "A TAG RFID deve estar entre 1 e 10.",
                "erro"
            )

            return redirect("/cadastrar")


        # ----------------------------------------------------
        # VERIFICA SE A TAG JÁ ESTÁ CADASTRADA
        # ----------------------------------------------------

        if Moto.query.filter_by(
            tag_rfid=tag
        ).first():

            flash(
                "Esta TAG RFID já está cadastrada.",
                "erro"
            )

            return redirect("/cadastrar")


        # ----------------------------------------------------
        # CAMPOS OBRIGATÓRIOS
        # ----------------------------------------------------

        if not chassi:

            flash(
                "Informe o chassi da moto.",
                "erro"
            )

            return redirect("/cadastrar")


        if not modelo_digitado:

            flash(
                "Informe o modelo da moto.",
                "erro"
            )

            return redirect("/cadastrar")


        if not cor_digitada:

            flash(
                "Informe a cor da moto.",
                "erro"
            )

            return redirect("/cadastrar")


        if not ano_digitado:

            flash(
                "Informe o ano da moto.",
                "erro"
            )

            return redirect("/cadastrar")


        # ----------------------------------------------------
        # PROCURA O CHASSI NO ARQUIVO IMPORTADO
        # ----------------------------------------------------

        moto_importada = MotoImportada.query.filter_by(
            chassi=chassi
        ).first()


        if moto_importada is None:

            flash(
                "Chassi não encontrado no arquivo importado.",
                "erro"
            )

            return redirect("/cadastrar")


        # ----------------------------------------------------
        # VERIFICA SE O CHASSI JÁ FOI CADASTRADO
        # ----------------------------------------------------

        if Moto.query.filter_by(
            chassi=moto_importada.chassi
        ).first():

            flash(
                "Esta moto já foi cadastrada.",
                "erro"
            )

            return redirect("/cadastrar")


        # ----------------------------------------------------
        # COMPARA MODELO
        # ----------------------------------------------------

        modelo_formulario = (
            modelo_digitado
            .strip()
            .upper()
        )

        modelo_arquivo = (
            moto_importada.modelo
            .strip()
            .upper()
        )


        if modelo_formulario != modelo_arquivo:

            flash(
                f"Modelo incorreto para este chassi. "
                f"Modelo correto: {moto_importada.modelo}",
                "erro"
            )

            return redirect("/cadastrar")


        # ----------------------------------------------------
        # COMPARA COR
        # ----------------------------------------------------

        cor_formulario = (
            cor_digitada
            .strip()
            .upper()
        )

        cor_arquivo = (
            moto_importada.cor
            .strip()
            .upper()
        )


        if cor_formulario != cor_arquivo:

            flash(
                f"Cor incorreta para este chassi. "
                f"Cor correta: {moto_importada.cor}",
                "erro"
            )

            return redirect("/cadastrar")


        # ----------------------------------------------------
        # COMPARA ANO
        # ----------------------------------------------------

        ano_formulario = (
            ano_digitado
            .strip()
        )

        ano_arquivo = (
            moto_importada.ano
            .strip()
        )


        if ano_formulario != ano_arquivo:

            flash(
                f"Ano incorreto para este chassi. "
                f"Ano correto: {moto_importada.ano}",
                "erro"
            )

            return redirect("/cadastrar")


        # ----------------------------------------------------
        # CADASTRA A MOTO
        # ----------------------------------------------------

        nova = Moto(

            tag_rfid=tag,

            chassi=moto_importada.chassi,

            modelo=moto_importada.modelo,

            cor=moto_importada.cor,

            ano=moto_importada.ano,

            placa=placa,

            montador=montador,

            local_atual="MONTAGEM",

            status="AGUARDANDO"

        )


        db.session.add(nova)

        db.session.commit()


        flash(
            "Moto cadastrada com sucesso!",
            "sucesso"
        )

        return redirect("/")


    return render_template(
        "cadastrar.html"
    )


# ============================================================
# SIMULAR LEITURA RFID
# ============================================================

@app.route("/ler")
def ler():

    tag = request.args.get(
        "tag",
        type=int
    )


    if tag is None:

        return "Informe uma TAG."


    moto = Moto.query.filter_by(
        tag_rfid=tag
    ).first()


    if moto is None:

        return f"TAG {tag} não cadastrada."


    setor = "MONTAGEM"

    moto.local_atual = setor


    leitura = Leitura(

        tag_rfid=str(tag),

        setor=setor,

        data_hora=datetime.now().strftime(
            "%d/%m/%Y %H:%M:%S"
        )

    )


    db.session.add(leitura)

    db.session.commit()


    exportar_txt(moto)


    return render_template(
        "resultado.html",
        moto=moto
    )


# ============================================================
# DESVINCULAR MOTO E LIBERAR TAG
# ============================================================

@app.route("/desvincular/<int:id>")
def desvincular(id):

    moto = Moto.query.get_or_404(id)


    # --------------------------------------------------------
    # SALVA NO HISTÓRICO
    # --------------------------------------------------------

    historico = HistoricoMoto(

        tag_rfid=str(
            moto.tag_rfid
        ),

        chassi=moto.chassi,

        modelo=moto.modelo,

        cor=moto.cor,

        ano=moto.ano,

        placa=moto.placa,

        montador=moto.montador

    )


    db.session.add(historico)


    # --------------------------------------------------------
    # REGISTRA A ENTREGA
    # --------------------------------------------------------

    leitura = Leitura(

        tag_rfid=str(
            moto.tag_rfid
        ),

        setor="MOTO ENTREGUE",

        data_hora=datetime.now().strftime(
            "%d/%m/%Y %H:%M:%S"
        )

    )


    db.session.add(leitura)


    # --------------------------------------------------------
    # REMOVE DO CADASTRO ATIVO
    # A TAG FICA DISPONÍVEL PARA NOVA MOTO
    # --------------------------------------------------------

    db.session.delete(moto)

    db.session.commit()


    flash(
        "Moto desvinculada e TAG liberada com sucesso!",
        "sucesso"
    )


    return redirect("/")


# ============================================================
# LEITURAS
# ============================================================

@app.route("/leituras")
def leituras():

    lista = Leitura.query.order_by(
        Leitura.id.desc()
    ).all()


    return render_template(
        "leituras.html",
        leituras=lista
    )


# ============================================================
# IMPORTAR TXT / CSV
# ============================================================

@app.route("/importar", methods=["GET", "POST"])
def importar():

    if request.method == "POST":

        arquivo = request.files.get(
            "arquivo"
        )


        # ----------------------------------------------------
        # VERIFICA ARQUIVO
        # ----------------------------------------------------

        if (
            arquivo is None
            or arquivo.filename == ""
        ):

            flash(
                "Nenhum arquivo foi selecionado.",
                "erro"
            )

            return redirect("/importar")


        try:

            # ------------------------------------------------
            # LÊ O ARQUIVO
            # ------------------------------------------------

            conteudo = arquivo.read().decode(
                "utf-8-sig"
            )


            if not conteudo.strip():

                flash(
                    "O arquivo está vazio.",
                    "erro"
                )

                return redirect("/importar")


            # ------------------------------------------------
            # DETECTA SEPARADOR
            # ------------------------------------------------

            primeira_linha = (
                conteudo
                .splitlines()[0]
            )


            if ";" in primeira_linha:

                separador = ";"

            else:

                separador = ","


            leitor = csv.DictReader(

                io.StringIO(
                    conteudo
                ),

                delimiter=separador

            )


            # ------------------------------------------------
            # NORMALIZA CABEÇALHOS
            # ------------------------------------------------

            if leitor.fieldnames:

                leitor.fieldnames = [

                    coluna.strip().upper()

                    for coluna
                    in leitor.fieldnames

                ]


            obrigatorias = [
                "CHASSI",
                "MODELO",
                "COR",
                "ANO"
            ]


            encontradas = (
                leitor.fieldnames or []
            )


            faltando = [

                coluna

                for coluna in obrigatorias

                if coluna not in encontradas

            ]


            if faltando:

                flash(
                    "Erro no arquivo. "
                    "Colunas não encontradas: "
                    + ", ".join(faltando),
                    "erro"
                )

                return redirect("/importar")


            importadas = 0

            duplicadas = 0


            # ------------------------------------------------
            # PROCESSA ARQUIVO
            # ------------------------------------------------

            for linha in leitor:

                chassi = (
                    linha.get(
                        "CHASSI",
                        ""
                    ) or ""
                ).strip()


                modelo = (
                    linha.get(
                        "MODELO",
                        ""
                    ) or ""
                ).strip()


                cor = (
                    linha.get(
                        "COR",
                        ""
                    ) or ""
                ).strip()


                ano = (
                    linha.get(
                        "ANO",
                        ""
                    ) or ""
                ).strip()


                if not chassi:

                    continue


                # --------------------------------------------
                # CHASSI JÁ EXISTE?
                # --------------------------------------------

                existente = MotoImportada.query.filter_by(
                    chassi=chassi
                ).first()


                if existente:

                    duplicadas += 1

                    continue


                # --------------------------------------------
                # NOVO REGISTRO
                # --------------------------------------------

                nova = MotoImportada(

                    chassi=chassi,

                    modelo=modelo,

                    cor=cor,

                    ano=ano

                )


                db.session.add(nova)

                importadas += 1


            db.session.commit()


            flash(

                f"Arquivo importado com sucesso! "
                f"{importadas} novo(s) registro(s) adicionado(s). "
                f"{duplicadas} registro(s) já existia(m).",

                "sucesso"

            )


            return redirect(
                "/importar"
            )


        except Exception as erro:

            db.session.rollback()


            flash(

                f"Erro ao importar o arquivo: {erro}",

                "erro"

            )


            return redirect(
                "/importar"
            )


    return render_template(
        "importar.html"
    )


# ============================================================
# RELATÓRIOS
# ============================================================

@app.route("/relatorios")
def relatorios():

    motos = Moto.query.order_by(
        Moto.id.desc()
    ).all()


    return render_template(
        "relatorios.html",
        motos=motos
    )


# ============================================================
# EXPORTAR RELATÓRIO PARA EXCEL
# ============================================================

@app.route("/relatorios/exportar")
def exportar_relatorio():

    data_inicial_texto = request.args.get(
        "data_inicial",
        ""
    ).strip()


    data_final_texto = request.args.get(
        "data_final",
        ""
    ).strip()


    # --------------------------------------------------------
    # VERIFICA DATAS
    # --------------------------------------------------------

    if (
        not data_inicial_texto
        or not data_final_texto
    ):

        flash(
            "Informe a data inicial e a data final.",
            "erro"
        )

        return redirect(
            "/relatorios"
        )


    # --------------------------------------------------------
    # CONVERTE DATAS
    # --------------------------------------------------------

    try:

        data_inicial = datetime.strptime(
            data_inicial_texto,
            "%Y-%m-%d"
        )


        data_final = datetime.strptime(
            data_final_texto,
            "%Y-%m-%d"
        ).replace(

            hour=23,
            minute=59,
            second=59

        )

    except ValueError:

        flash(
            "Data inválida.",
            "erro"
        )

        return redirect(
            "/relatorios"
        )


    # --------------------------------------------------------
    # VALIDA PERÍODO
    # --------------------------------------------------------

    if data_inicial > data_final:

        flash(
            "A data inicial não pode ser maior que a data final.",
            "erro"
        )

        return redirect(
            "/relatorios"
        )


    # --------------------------------------------------------
    # BUSCA MOTOS DO PERÍODO
    # --------------------------------------------------------

    motos = Moto.query.filter(

        Moto.data_cadastro >= data_inicial,

        Moto.data_cadastro <= data_final

    ).order_by(

        Moto.data_cadastro.asc()

    ).all()


    # ========================================================
    # CRIA EXCEL
    # ========================================================

    wb = Workbook()

    ws = wb.active

    ws.title = "Relatório de Motos"


    # --------------------------------------------------------
    # TÍTULO
    # --------------------------------------------------------

    ws["A1"] = "RELATÓRIO DE MOTOS CADASTRADAS"

    ws["A1"].font = Font(
        bold=True,
        size=16
    )

    ws.merge_cells(
        "A1:J1"
    )


    # --------------------------------------------------------
    # PERÍODO
    # --------------------------------------------------------

    ws["A2"] = "PERÍODO"

    ws["A2"].font = Font(
        bold=True
    )


    ws["B2"] = (

        data_inicial.strftime(
            "%d/%m/%Y"
        )

        + " até "

        + data_final.strftime(
            "%d/%m/%Y"
        )

    )


    # --------------------------------------------------------
    # CABEÇALHO
    # --------------------------------------------------------

    cabecalho = [

        "DATA DO CADASTRO",
        "TAG RFID",
        "CHASSI",
        "MODELO",
        "COR",
        "ANO",
        "PLACA",
        "MONTADOR",
        "LOCAL ATUAL",
        "STATUS"

    ]


    linha_cabecalho = 4


    for coluna, titulo in enumerate(
        cabecalho,
        start=1
    ):

        celula = ws.cell(
            row=linha_cabecalho,
            column=coluna
        )

        celula.value = titulo

        celula.font = Font(
            bold=True
        )

        celula.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )


    # --------------------------------------------------------
    # DADOS
    # --------------------------------------------------------

    linha_excel = 5


    for moto in motos:

        data_cadastro = ""


        if moto.data_cadastro:

            data_cadastro = moto.data_cadastro.strftime(
                "%d/%m/%Y %H:%M:%S"
            )


        valores = [

            data_cadastro,

            moto.tag_rfid,

            moto.chassi,

            moto.modelo,

            moto.cor,

            moto.ano,

            moto.placa or "",

            moto.montador or "",

            moto.local_atual or "",

            moto.status or ""

        ]


        for coluna, valor in enumerate(
            valores,
            start=1
        ):

            celula = ws.cell(
                row=linha_excel,
                column=coluna
            )

            celula.value = valor

            celula.alignment = Alignment(
                vertical="center"
            )


        linha_excel += 1


    # --------------------------------------------------------
    # LARGURA DAS COLUNAS
    # --------------------------------------------------------

    larguras = {

        "A": 22,
        "B": 12,
        "C": 25,
        "D": 28,
        "E": 22,
        "F": 15,
        "G": 15,
        "H": 25,
        "I": 20,
        "J": 20

    }


    for coluna, largura in larguras.items():

        ws.column_dimensions[
            coluna
        ].width = largura


    # --------------------------------------------------------
    # CONGELA CABEÇALHO
    # --------------------------------------------------------

    ws.freeze_panes = "A5"


    # --------------------------------------------------------
    # GERA ARQUIVO NA MEMÓRIA
    # --------------------------------------------------------

    arquivo = BytesIO()

    wb.save(
        arquivo
    )

    arquivo.seek(0)


    # --------------------------------------------------------
    # NOME DO ARQUIVO
    # --------------------------------------------------------

    nome_arquivo = (

        "relatorio_motos_"

        + data_inicial.strftime(
            "%d-%m-%Y"
        )

        + "_a_"

        + data_final.strftime(
            "%d-%m-%Y"
        )

        + ".xlsx"

    )


    # --------------------------------------------------------
    # DOWNLOAD
    # --------------------------------------------------------

    return send_file(

        arquivo,

        as_attachment=True,

        download_name=nome_arquivo,

        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )

    )


# ============================================================
# SIMULAR TAG ESPECÍFICA
# ============================================================

@app.route("/ler/<int:tag>")
def ler_manual(tag):

    if tag < 1 or tag > 10:

        return "TAG inválida."


    moto = Moto.query.filter_by(
        tag_rfid=tag
    ).first()


    if moto is None:

        return f"TAG {tag} não cadastrada."


    moto.local_atual = "MONTAGEM"


    if hasattr(
        moto,
        "status"
    ):

        moto.status = "EM PROCESSO"


    leitura = Leitura(

        tag_rfid=str(tag),

        setor="MONTAGEM",

        data_hora=datetime.now().strftime(
            "%d/%m/%Y %H:%M:%S"
        )

    )


    db.session.add(leitura)

    db.session.commit()


    exportar_txt(moto)


    return render_template(
        "resultado.html",
        moto=moto
    )


# ============================================================
# INICIA O SERVIDOR
# ============================================================

if __name__ == "__main__":

    print()
    print("==============================================")
    print("      SISTEMA RFID DE RASTREAMENTO")
    print("==============================================")
    print()
    print("HTML: pasta raiz")
    print("CSS : pasta raiz")
    print()
    print("Acesso local:")
    print("http://127.0.0.1:5000")
    print()
    print("Acesso pela rede:")
    print("http://IP-DO-COMPUTADOR:5000")
    print()
    print("==============================================")
    print()


    app.run(
        host="0.0.0.0",
        port=5000,
        debug=True
    )